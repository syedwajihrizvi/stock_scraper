from openpyxl import Workbook
from statements import indicators


class ComparitiveSheet:
    __column_letters = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']

    def __init__(self, company):
        self.company = company
        self.__wb = Workbook()
        self.general_worksheet = self.__wb.create_sheet('General')
        self.cfs_worksheet = self.__wb.create_sheet('Cash Flow')
        self.income_worksheet = self.__wb.create_sheet('Income')
        self.balance_worksheet = self.__wb.create_sheet('Balance Sheet')
        self.key_stats_worksheet = self.__wb.create_sheet('Key Stats')
        self.fin_data_worksheet = self.__wb.create_sheet('Financials')
        self.sum_detail_worksheet = self.__wb.create_sheet('Summary')

    def changeColumnWidth(self, width):
        for sheet in self.__wb:
            sheet.column_dimensions['A'].width = width
            for column in self.__column_letters:
                sheet.column_dimensions[column].width = width

    def create_keys(self, keys, sheet_name):
        cell_number = 1
        sheet = self.__wb[sheet_name]
        sheet['A1'] = 'Key'
        for key in keys:
            cell_number = cell_number + 1
            sheet[f'A{cell_number}'] = key['label']

    def create_company_names(self, companies):
        for sheet in self.__wb:
            for index in range(len(companies)):
                column = self.__column_letters[index]
                sheet[f'{column}1'] = companies[index].company.capitalize()

    def save(self):
        self.__wb.save("/Users/wajihrizvi/Desktop/python-projects/test.xlsx")

    def upload_data(self, data, column, sheet_name):
        column_letter = self.__column_letters[column]
        cell_number = 1
        sheet = self.__wb[sheet_name]
        for key in data:
            cell_number = cell_number + 1
            sheet[f'{column_letter}{cell_number}'] = data[key].get('value')
